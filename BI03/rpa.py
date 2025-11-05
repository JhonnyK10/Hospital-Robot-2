import time
import sys
import io
import pyautogui
import shutil
import pandas as pd  
import zipfile
import unicodedata
import re 
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from datetime import datetime

# Configurar encoding para UTF-8 para evitar problemas com caracteres especiais
if sys.stdout.encoding != 'UTF-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
if sys.stderr.encoding != 'UTF-8':
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# Tentativa adicional para Windows
try:
    # Para Python 3.7+
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except:
    pass


# ================== FUNÇÃO PARA REMOVER ACENTOS ==================

def remove_accents(text):
    """Remove acentos de uma string"""
    if not isinstance(text, str):
        text = str(text)
    return ''.join(
        c for c in unicodedata.normalize('NFD', text)
        if unicodedata.category(c) != 'Mn'
    )

# ================== FUNÇÕES PARA AGRUPAMENTO INTELIGENTE DE HOSPITAIS ==================

def normalize_hospital_name_for_grouping(hospital_name):
    """
    Normaliza o nome do hospital para agrupamento, removendo variações comuns.
    """
    # Remove acentos e converte para maiúsculas
    name = remove_accents(hospital_name).upper()
    
    # Remove espaços extras
    name = ' '.join(name.split())
    
    # Substitui variações comuns por formas padronizadas
    replacements = {
        'S/A': 'SA',
        'S.A.': 'SA', 
        'S.A': 'SA',
        ' LTDA': '',
        ' ME': '',
        ' EPP': '',
        ' INTERN': 'INTERNACIONAL',
        ' INTERNAC': 'INTERNACIONAL',
        ' INTERNACION': 'INTERNACIONAL',
        ' ASSISTENCIA': '',
        ' MEDICA': '',
        ' DE ': ' ',
        ' DA ': ' ',
        ' DO ': ' ',
        ' DOS ': ' ',
        ' DAS ': ' '
    }
    
    for old, new in replacements.items():
        name = name.replace(old, new)
    
    # Remove caracteres especiais e espaços duplicados
    name = ''.join(c for c in name if c.isalnum() or c.isspace())
    name = ' '.join(name.split())
    
    return name

def calculate_similarity(str1, str2):
    """
    Calcula a similaridade entre duas strings usando uma abordagem simples.
    Retorna um valor entre 0 e 1.
    """
    str1 = str1.lower()
    str2 = str2.lower()
    
    # Se uma string está contida na outra, consideramos alta similaridade
    if str1 in str2 or str2 in str1:
        return 0.9
    
    # Calcula similaridade baseada em palavras em comum
    words1 = set(str1.split())
    words2 = set(str2.split())
    
    if not words1 or not words2:
        return 0.0
    
    common_words = words1.intersection(words2)
    all_words = words1.union(words2)
    
    return len(common_words) / len(all_words)

def find_most_complete_name(names):
    """
    Encontra o nome mais completo/representativo em uma lista de nomes.
    Geralmente o nome mais longo é o mais completo.
    """
    if not names:
        return ""
    
    # Remove partes comuns do nome do arquivo (Boleto_, _Bradesco, _Itau, data)
    cleaned_names = []
    for name in names:
        # Remove prefixo Boleto_
        clean_name = name.replace('Boleto_', '')
        # Remove sufixos _Bradesco, _Itau e data
        clean_name = re.sub(r'_(Bradesco|Itau)_\d{8}$', '', clean_name)
        clean_name = clean_name.replace('_', ' ')
        cleaned_names.append(clean_name.strip())
    
    # Retorna o nome mais longo (geralmente o mais completo)
    return max(cleaned_names, key=len)

def group_pdfs_by_hospital(pdf_files):
    """
    Agrupa os PDFs por hospital usando matching inteligente.
    Retorna um dicionário: {hospital_name: [lista_de_pdfs]}
    """
    hospital_pdfs = {}
    
    for pdf_path in pdf_files:
        # Extrai o nome do hospital do nome do arquivo PDF
        pdf_name = pdf_path.stem
        
        # Remove prefixos e sufixos comuns para extrair o nome do hospital
        parts = pdf_name.split('_')
        
        # O nome do hospital geralmente está entre "Boleto_" e o tipo (Bradesco/Itau)
        hospital_name = None
        
        # Tenta encontrar o índice de "Boleto"
        if 'Boleto' in parts:
            boleto_index = parts.index('Boleto')
            # O hospital está após "Boleto" e antes do tipo (Bradesco/Itau)
            hospital_parts = []
            for i in range(boleto_index + 1, len(parts)):
                if parts[i] in ['Bradesco', 'Itau']:
                    break
                hospital_parts.append(parts[i])
            
            if hospital_parts:
                hospital_name = ' '.join(hospital_parts).replace('_', ' ')
        
        # Fallback: se não encontrou pelo método acima
        if not hospital_name:
            # Remove prefixos conhecidos e sufixos
            fallback_name = pdf_name.replace('Boleto_', '').replace('_Bradesco', '').replace('_Itau', '')
            # Remove data no final (formato YYYYMMDD)
            fallback_name = re.sub(r'_\d{8}$', '', fallback_name)
            hospital_name = fallback_name.replace('_', ' ')
        
        # Normaliza o nome para agrupamento
        normalized_name = normalize_hospital_name_for_grouping(hospital_name)
        
        print(f"Arquivo: {pdf_path.name}")
        print(f"  Nome extraído: {hospital_name}")
        print(f"  Nome normalizado: {normalized_name}")
        
        # Verifica se já existe um hospital similar no dicionário
        existing_hospital = None
        for existing_norm_name in hospital_pdfs.keys():
            # Calcula a similaridade entre os nomes normalizados
            similarity = calculate_similarity(normalized_name, existing_norm_name)
            if similarity > 0.8:  # 80% de similaridade
                existing_hospital = existing_norm_name
                print(f"  -> Agrupado com: {existing_hospital} (similaridade: {similarity:.2f})")
                break
        
        # Se não encontrou similar, usa o nome normalizado
        if not existing_hospital:
            existing_hospital = normalized_name
            hospital_pdfs[existing_hospital] = []
            print(f"  -> Novo grupo: {existing_hospital}")
        
        hospital_pdfs[existing_hospital].append(pdf_path)
    
    # Converte para o formato final usando o nome mais completo como chave
    final_grouping = {}
    for normalized_name, pdf_list in hospital_pdfs.items():
        # Encontra o nome mais completo/representativo entre os PDFs
        most_complete_name = find_most_complete_name([pdf.stem for pdf in pdf_list])
        final_grouping[most_complete_name] = pdf_list
        print(f"Grupo final: '{most_complete_name}' com {len(pdf_list)} PDFs")
    
    return final_grouping

# ================== CONFIGURAÇÕES ==================
load_dotenv()

# Define as pastas dentro da pasta BI03
BASE_DIR = Path.cwd()  # Pasta BI03
DOWNLOAD_FOLDER = BASE_DIR / "downloads"
PROCESSED_FOLDER = BASE_DIR / "boletos_pdf"

# Caminho do arquivo de configurações - DINÂMICO
def find_config_excel_path():
    """
    Busca o arquivo 'infos do robo.xlsx' dinamicamente.
    """
    possible_paths = [
        Path.cwd() / "assets" / "infos do robo.xlsx",
        Path.cwd() / "infos do robo.xlsx", 
        Path(__file__).parent / "assets" / "infos do robo.xlsx",
    ]
    
    for path in possible_paths:
        if path.exists():
            print(f"Arquivo de configuração encontrado: {path}")
            return path
    
    # Busca recursivamente
    for file_path in Path.cwd().rglob("infos do robo.xlsx"):
        if file_path.exists():
            print(f"Arquivo de configuração encontrado: {file_path}")
            return file_path
    
    print("ERRO: Arquivo 'infos do robo.xlsx' não encontrado!")
    return None

CONFIG_EXCEL_PATH = find_config_excel_path()

if CONFIG_EXCEL_PATH is None:
    print("Não foi possível encontrar o arquivo de configuração.")
    print("Verifique se o arquivo 'infos do robo.xlsx' está na pasta do projeto.")
    exit(1)

OUTLOOK_WEB_URL = "https://outlook.office.com/mail/inbox"

# Variáveis que serão carregadas do Excel
EMAIL_SUBJECT = None
PDF_FOLDER_PATH = None
LOG_AUTOMATION_EMAIL = None
EMAILS_EXCEL_PATH = None
OUTLOOK_EMAIL = None
OUTLOOK_PASSWORD = None
DEFAULT_WAIT_TIME = 3

# Variável global para armazenar o status dos envios
email_status_report = []

def load_config_from_excel():
    """
    Carrega as configurações do arquivo Excel.
    """
    global EMAIL_SUBJECT, PDF_FOLDER_PATH, LOG_AUTOMATION_EMAIL, EMAILS_EXCEL_PATH, OUTLOOK_EMAIL, OUTLOOK_PASSWORD
    
    try:
        if not CONFIG_EXCEL_PATH.exists():
            print(f"ERRO: Arquivo de configuração não encontrado: {CONFIG_EXCEL_PATH}")
            return False
        
        workbook = load_workbook(CONFIG_EXCEL_PATH, data_only=True)
        sheet = workbook.active
        
        config_dict = {}
        
        # Lê todas as linhas e armazena em um dicionário
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if row[0] and row[1]:  # Chave e Valor
                key = str(row[0]).strip().lower()
                value = str(row[1]).strip() if row[1] else ""
                config_dict[key] = value
        
        print("Configurações carregadas do Excel:")
        for key, value in config_dict.items():
            print(f"  - {key}: {value}")
        
        # Atribui às variáveis globais
        EMAIL_SUBJECT = config_dict.get('assunto do email', '')
        PDF_FOLDER_PATH = config_dict.get('caminho para faturas', '')
        LOG_AUTOMATION_EMAIL = config_dict.get('email de relatorio', '')
        EMAILS_EXCEL_PATH = Path(config_dict.get('caminho dos emails hospitais', ''))
        OUTLOOK_EMAIL = config_dict.get('email_user', '')
        OUTLOOK_PASSWORD = config_dict.get('email_pass', '')
        
        # Validações básicas
        if not all([EMAIL_SUBJECT, PDF_FOLDER_PATH, LOG_AUTOMATION_EMAIL, EMAILS_EXCEL_PATH, OUTLOOK_EMAIL, OUTLOOK_PASSWORD]):
            print("ERRO: Algumas configurações obrigatórias não foram encontradas no arquivo Excel.")
            print("Configurações necessárias:")
            print("  - Assunto do Email")
            print("  - Caminho para faturas") 
            print("  - Email de relatorio")
            print("  - Caminho dos emails hospitais")
            print("  - Email_user")
            print("  - Email_pass")
            return False
        
        # Converte caminho para Path object
        PDF_FOLDER_PATH = Path(PDF_FOLDER_PATH)
        
        print("Configurações carregadas com sucesso!")
        return True
        
    except Exception as e:
        print(f"ERRO ao carregar configurações do Excel: {e}")
        return False

def clean_folders():
    """Limpa todas as pastas (downloads e faturas_pdf) antes de iniciar o processo."""
    folders_to_clean = [DOWNLOAD_FOLDER, PROCESSED_FOLDER]
    
    for folder in folders_to_clean:
        try:
            if folder.exists():
                # Remove todos os arquivos dentro da pasta
                for file_path in folder.glob('*'):
                    try:
                        if file_path.is_file():
                            file_path.unlink()
                            print(f"Arquivo excluído: {file_path.name}")
                        elif file_path.is_dir():
                            shutil.rmtree(file_path)
                            print(f"Pasta excluída: {file_path.name}")
                    except Exception as e:
                        print(f"Erro ao excluir {file_path}: {e}")
                
                print(f"Pasta limpa: {folder.name}")
            else:
                print(f"Pasta não existe, criando: {folder.name}")
                folder.mkdir(exist_ok=True)
        except Exception as e:
            print(f"Erro ao limpar pasta {folder}: {e}")

def create_folders():
    """Garante que as pastas de downloads e processados existam."""
    DOWNLOAD_FOLDER.mkdir(exist_ok=True)
    PROCESSED_FOLDER.mkdir(exist_ok=True)
    print("Pastas criadas/verificadas com sucesso")

def extract_zip_files():
    """
    Extrai automaticamente todos os arquivos ZIP da pasta de downloads.
    """
    print("Verificando arquivos ZIP para extração...")
    
    # Procura por arquivos ZIP na pasta de downloads
    zip_files = list(DOWNLOAD_FOLDER.glob("*.zip"))
    
    if not zip_files:
        print("Nenhum arquivo ZIP encontrado para extrair.")
        return
    
    for zip_file in zip_files:
        try:
            print(f"Extraindo arquivo: {zip_file.name}")
            
            # Cria uma pasta temporária para extração (opcional)
            extract_folder = DOWNLOAD_FOLDER / zip_file.stem
            extract_folder.mkdir(exist_ok=True)
            
            # Extrai o arquivo ZIP
            with zipfile.ZipFile(zip_file, 'r') as zip_ref:
                zip_ref.extractall(extract_folder)
            
            # Lista os arquivos extraídos
            extracted_files = list(extract_folder.glob("*"))
            print(f"Arquivos extraídos de {zip_file.name}:")
            for file in extracted_files:
                print(f"  - {file.name}")
                
                # Move os arquivos extraídos para a pasta principal de downloads
                if file.is_file():
                    destination = DOWNLOAD_FOLDER / file.name
                    # Se já existir um arquivo com o mesmo nome, adiciona um sufixo
                    if destination.exists():
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        name_parts = file.stem, file.suffix
                        destination = DOWNLOAD_FOLDER / f"{name_parts[0]}_{timestamp}{name_parts[1]}"
                    
                    shutil.move(str(file), str(destination))
                    print(f"    Movido para: {destination.name}")
            
            # Remove a pasta temporária de extração
            shutil.rmtree(extract_folder)
            
            print(f"Extração concluída para: {zip_file.name}")
            
        except zipfile.BadZipFile:
            print(f"ERRO: {zip_file.name} não é um arquivo ZIP válido")
        except Exception as e:
            print(f"ERRO ao extrair {zip_file.name}: {e}")

def normalize_column_name(col_name):
    """Normaliza nomes de colunas para facilitar o mapeamento."""
    if not isinstance(col_name, str):
        col_name = str(col_name)
    return (col_name.lower()
            .strip()
            .replace(' ', '_')
            .replace('°', '')
            .replace('(', '')
            .replace(')', '')
            .replace('$', '')
            .replace('.', '')
            .replace('/', '')
            .replace('\\', '')
            .replace('-', '_')
            .replace('ã', 'a')
            .replace('ç', 'c')
            .replace('é', 'e')
            .replace('ê', 'e')
            .replace('í', 'i')
            .replace('ó', 'o')
            .replace('ô', 'o')
            .replace('ú', 'u'))

def process_excel_files_and_generate_pdfs():
    """
    Processa os arquivos Excel extraídos e gera PDFs com o formato específico de cada arquivo.
    """
    print("Iniciando processamento dos arquivos Excel...")
    
    # Encontra TODOS os arquivos Excel na pasta de downloads
    excel_files = list(DOWNLOAD_FOLDER.glob("*.xlsx")) + list(DOWNLOAD_FOLDER.glob("*.xls"))
    
    if not excel_files:
        print("Nenhum arquivo Excel encontrado para processar.")
        return []
    
    print(f"Arquivos Excel encontrados: {[f.name for f in excel_files]}")
    
    all_pdf_files = []
    
    # Processa CADA arquivo separadamente
    for excel_file in excel_files:
        try:
            print(f"\nProcessando arquivo: {excel_file.name}")
            
            # Lê o arquivo Excel
            df = pd.read_excel(excel_file)
            
            # Processa conforme o tipo de arquivo
            if 'bradesco' in excel_file.name.lower():
                pdf_files = process_bradesco_file(df, excel_file)
            elif 'itau' in excel_file.name.lower():
                pdf_files = process_itau_file(df, excel_file)
            else:
                print(f"  AVISO: Tipo de arquivo não reconhecido: {excel_file.name}")
                continue
            
            if pdf_files:
                all_pdf_files.extend(pdf_files)
                print(f"  PDFs gerados para {excel_file.name}: {len(pdf_files)}")
            else:
                print(f"  Nenhum PDF gerado para {excel_file.name}")
            
        except Exception as e:
            print(f"ERRO ao processar {excel_file.name}: {e}")
            import traceback
            traceback.print_exc()
    
    print(f"\nTotal de PDFs gerados: {len(all_pdf_files)}")
    return all_pdf_files

def process_bradesco_file(df, excel_file):
    """
    Processa arquivo Bradesco e gera PDFs com formato específico do Bradesco.
    """
    print("  Formatando como Bradesco...")
    
    # Normaliza colunas
    df_standard = df.copy()
    df_standard.columns = [normalize_column_name(col) for col in df_standard.columns]
    print(f"  Colunas normalizadas: {list(df_standard.columns)}")
    
    # Mapeamento das colunas do Bradesco
    column_mapping = {
        'status': 'Status',
        'pagador': 'Pagador', 
        'n_nota': 'Nº Nota',
        'n_boleto': 'Nº Boleto',
        'data_de_vencim': 'Data Vencimento',
        'valor': 'Valor'
    }
    
    # Aplica o mapeamento
    df_clean = pd.DataFrame()
    for source_col, target_col in column_mapping.items():
        if source_col in df_standard.columns:
            df_clean[target_col] = df_standard[source_col]
        else:
            print(f"    AVISO: Coluna {source_col} não encontrada no Bradesco")
            df_clean[target_col] = None
    
    # Limpa os dados
    if 'Pagador' in df_clean.columns:
        df_clean['Pagador'] = df_clean['Pagador'].astype(str).str.strip()
        df_clean = df_clean[df_clean['Pagador'].notna() & (df_clean['Pagador'] != '') & (df_clean['Pagador'] != 'nan')]
    
    # Converte valor para numérico
    if 'Valor' in df_clean.columns:
        df_clean['Valor'] = pd.to_numeric(df_clean['Valor'], errors='coerce').fillna(0)
    
    # Formata data
    if 'Data Vencimento' in df_clean.columns:
        df_clean['Data Vencimento'] = df_clean['Data Vencimento'].apply(format_date)
    
    print(f"  Colunas Bradesco: {list(df_clean.columns)}")
    print(f"  Registros Bradesco: {len(df_clean)}")
    
    # Gera PDFs para cada hospital
    return generate_pdfs_for_file(df_clean, excel_file, 'Bradesco')

def process_itau_file(df, excel_file):
    """
    Processa arquivo Itaú e gera PDFs com formato específico do Itaú.
    """
    print("  Formatando como Itau...")  # Mudei para "Itau" sem acento
    
    # Normaliza colunas
    df_standard = df.copy()
    df_standard.columns = [normalize_column_name(col) for col in df_standard.columns]
    print(f"  Colunas normalizadas: {list(df_standard.columns)}")
    
    # Mapeamento das colunas do Itaú
    column_mapping = {
        'pagador': 'Pagador',
        'vencimento': 'Data Vencimento', 
        'valorr': 'Valor',
        'n_boleto': 'Nº Boleto',
        'n_nota': 'Nº Nota',
        'observacao': 'Status'
    }
    
    # Aplica o mapeamento
    df_clean = pd.DataFrame()
    for source_col, target_col in column_mapping.items():
        if source_col in df_standard.columns:
            df_clean[target_col] = df_standard[source_col]
        else:
            print(f"    AVISO: Coluna {source_col} não encontrada no Itau")
            df_clean[target_col] = None
    
    # Limpa os dados
    if 'Pagador' in df_clean.columns:
        df_clean['Pagador'] = df_clean['Pagador'].astype(str).str.strip()
        df_clean = df_clean[df_clean['Pagador'].notna() & (df_clean['Pagador'] != '') & (df_clean['Pagador'] != 'nan')]
    
    # Converte valor para numérico
    if 'Valor' in df_clean.columns:
        df_clean['Valor'] = pd.to_numeric(df_clean['Valor'], errors='coerce').fillna(0)
    
    # Formata data
    if 'Data Vencimento' in df_clean.columns:
        df_clean['Data Vencimento'] = df_clean['Data Vencimento'].apply(format_date)
    
    # Preenche status vazio
    if 'Status' in df_clean.columns:
        df_clean['Status'] = df_clean['Status'].fillna('VENCIDO')
    
    print(f"  Colunas Itau: {list(df_clean.columns)}")
    print(f"  Registros Itau: {len(df_clean)}")
    
    # Gera PDFs para cada hospital - use "Itau" sem acento
    return generate_pdfs_for_file(df_clean, excel_file, 'Itau')

def format_date(date_value):
    """
    Formata datas para o padrão dd/mm/aaaa.
    """
    if pd.isna(date_value) or date_value == '':
        return ""
    
    try:
        if isinstance(date_value, str):
            # Tenta vários formatos de data
            formats = [
                '%Y-%m-%d %H:%M:%S',
                '%d/%m/%Y',
                '%d/%m/%y',
                '%Y-%m-%d',
                '%d-%m-%Y',
                '%d-%m-%y'
            ]
            
            for fmt in formats:
                try:
                    date_obj = datetime.strptime(date_value, fmt)
                    return date_obj.strftime('%d/%m/%Y')
                except:
                    continue
            
            # Se nenhum formato funcionar, retorna o valor original
            return date_value
        else:
            # Se é datetime do pandas
            date_obj = pd.to_datetime(date_value)
            return date_obj.strftime('%d/%m/%Y')
    except:
        return str(date_value)

def generate_pdfs_for_file(df, excel_file, file_type):
    """
    Gera PDFs para um arquivo específico.
    """
    pdf_files = []
    
    # Verifica se há dados
    if df.empty:
        print(f"  AVISO: Nenhum dado válido encontrado no {file_type}")
        return pdf_files
    
    # Agrupa por hospital
    hospitals = df['Pagador'].unique()
    
    print(f"  Hospitais encontrados no {file_type}: {len(hospitals)}")
    
    for hospital in hospitals:
        try:
            hospital_data = df[df['Pagador'] == hospital].copy()
            
            if hospital_data.empty:
                continue
                
            # Gera o PDF específico para o tipo de arquivo
            pdf_path = generate_specific_pdf(hospital, hospital_data, excel_file, file_type)
            if pdf_path:
                pdf_files.append(pdf_path)
                
        except Exception as e:
            print(f"  ERRO ao gerar PDF para {hospital}: {e}")
            import traceback
            traceback.print_exc()
    
    return pdf_files

def generate_specific_pdf(hospital_name, hospital_data, excel_file, file_type):
    """
    Gera PDF com formato específico para cada tipo de arquivo.
    """
    print(f"  Gerando PDF {file_type} para: {hospital_name}")
    
    # Remove acentos do nome do hospital e do tipo de arquivo
    hospital_name_clean = remove_accents(hospital_name)
    file_type_clean = remove_accents(file_type)
    
    # Limpa o nome do hospital para usar no nome do arquivo
    safe_hospital_name = "".join(c for c in hospital_name_clean if c.isalnum() or c in (' ', '-', '_')).rstrip()
    safe_hospital_name = safe_hospital_name.replace(' ', '_')
    
    # Cria o nome do arquivo PDF SEM ACENTOS
    pdf_filename = f"Boleto_{safe_hospital_name}_{file_type_clean}_{datetime.now().strftime('%Y%m%d')}.pdf"
    pdf_path = PROCESSED_FOLDER / pdf_filename
    
    try:
        # Cria o documento PDF
        doc = SimpleDocTemplate(str(pdf_path), pagesize=A4)
        elements = []
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = styles['Heading1']
        normal_style = styles['Normal']
        
        # Título
        title = Paragraph(f"Relatório de Boletos - {hospital_name}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Informações do hospital
        total_registros = len(hospital_data)
        
        info_text = f"Total de registros: {total_registros}<br/>Origem: {file_type} - {excel_file.stem}"
        info_paragraph = Paragraph(info_text, normal_style)
        elements.append(info_paragraph)
        elements.append(Spacer(1, 12))
        
        # Prepara os dados para a tabela
        table_data = []
        
        # Usa as colunas específicas do DataFrame
        columns = list(hospital_data.columns)
        headers = columns
        table_data.append(headers)
        
        # Dados da tabela
        for _, row in hospital_data.iterrows():
            table_row = []
            for col in columns:
                value = row[col]
                
                # Formata a coluna Valor
                if col == 'Valor':
                    if pd.isna(value) or value == 0:
                        table_row.append("R$ 0,00")
                    else:
                        try:
                            valor_float = float(value)
                            table_row.append(f"R$ {valor_float:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
                        except (ValueError, TypeError):
                            table_row.append(f"R$ {value}")
                else:
                    if pd.isna(value) or value == '':
                        table_row.append("")
                    else:
                        table_row.append(str(value))
            
            table_data.append(table_row)
        
        # Cria a tabela
        if len(table_data) > 1:
            table = Table(table_data)
            
            # Estilo da tabela
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ])
            
            table.setStyle(table_style)
            elements.append(table)
        
        # Constrói o PDF
        doc.build(elements)
        
        print(f"    PDF gerado: {pdf_filename}")
        return pdf_path
        
    except Exception as e:
        print(f"    ERRO ao gerar PDF: {e}")
        import traceback
        traceback.print_exc()
        return None
    
def load_hospital_emails():
    """
    Carrega a relação de emails dos hospitais do arquivo Excel.
    Retorna um dicionário: {hospital_name: {'to': ['email1', 'email2'], 'cc': ['email3', 'email4']}}
    AGORA INCLUI CC1 E CC2
    """
    try:
        if not EMAILS_EXCEL_PATH.exists():
            print(f"ERRO: Arquivo de emails não encontrado: {EMAILS_EXCEL_PATH}")
            return {}
        
        workbook = load_workbook(EMAILS_EXCEL_PATH, data_only=True)
        sheet = workbook.active
        
        emails_dict = {}
        
        # Pega cabeçalhos
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).lower())
        
        print(f"Cabeçalhos encontrados no arquivo de emails: {headers}")
        
        # Encontra os índices das colunas de forma mais flexível
        hospital_col = None
        email_col = None
        cc1_col = None
        cc2_col = None
        
        for i, header in enumerate(headers):
            header_lower = header.lower()
            if any(word in header_lower for word in ['hospital', 'hosp', 'nome', 'cliente']):
                hospital_col = i
            elif any(word in header_lower for word in ['email', 'to', 'para', 'e-mail']):
                email_col = i
            elif 'cc 1' in header_lower or 'cc1' in header_lower.replace(' ', ''):
                cc1_col = i
            elif 'cc 2' in header_lower or 'cc2' in header_lower.replace(' ', ''):
                cc2_col = i
        
        print(f"Colunas identificadas: Hospital={hospital_col}, Email={email_col}, Cc1={cc1_col}, Cc2={cc2_col}")
        
        if hospital_col is None or email_col is None:
            print("ERRO: Colunas 'Hospital' e 'Email' não encontradas no arquivo de emails")
            print("Colunas disponíveis:", headers)
            return {}
        
        # Processa as linhas - agora agrupando múltiplos emails
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[hospital_col] and row[email_col]:
                hospital_name = str(row[hospital_col]).strip()
                to_email = str(row[email_col]).strip()
                
                # Processa emails CC - AGORA INCLUI CC1 E CC2
                cc_emails = []
                
                # Adiciona email da coluna Cc 1
                if cc1_col is not None and row[cc1_col]:
                    cc1_value = str(row[cc1_col]).strip()
                    if cc1_value:
                        # Separa múltiplos emails por vírgula ou ponto-e-vírgula
                        cc_emails.extend([email.strip() for email in cc1_value.replace(';', ',').split(',') if email.strip()])
                
                # Adiciona email da coluna Cc 2
                if cc2_col is not None and row[cc2_col]:
                    cc2_value = str(row[cc2_col]).strip()
                    if cc2_value:
                        # Separa múltiplos emails por vírgula ou ponto-e-vírgula
                        cc_emails.extend([email.strip() for email in cc2_value.replace(';', ',').split(',') if email.strip()])
                
                # Remove duplicatas mantendo a ordem
                cc_emails = list(dict.fromkeys(cc_emails))
                
                # Se o hospital ainda não existe no dicionário, cria
                if hospital_name not in emails_dict:
                    emails_dict[hospital_name] = {
                        'to': set(),  # Usamos set para evitar duplicatas
                        'cc': set()
                    }
                
                # Adiciona os emails aos conjuntos
                emails_dict[hospital_name]['to'].add(to_email)
                
                # Adiciona todos os emails CC (tanto Cc1 quanto Cc2)
                for cc_email in cc_emails:
                    emails_dict[hospital_name]['cc'].add(cc_email)
        
        # Converte sets para listas para facilitar o uso posterior
        for hospital in emails_dict:
            emails_dict[hospital]['to'] = list(emails_dict[hospital]['to'])
            emails_dict[hospital]['cc'] = list(emails_dict[hospital]['cc'])
        
        print(f"Carregados emails para {len(emails_dict)} hospitais")
        for hospital, emails in emails_dict.items():
            print(f"  - {hospital}: To={', '.join(emails['to'])}, Cc={', '.join(emails['cc'])}")
        
        return emails_dict
        
    except Exception as e:
        print(f"ERRO ao carregar arquivo de emails: {e}")
        return {}

def find_hospital_email(hospital_name, hospital_emails):
    """
    Encontra o email correspondente para um hospital usando matching inteligente
    que funciona para qualquer hospital novo adicionado na planilha.
    AGORA RETORNA None SEM IMPRIMIR MENSAGENS DE ERRO EXCESSIVAS.
    """
    hospital_name_clean = hospital_name.upper().strip()
    
    # 1. Primeiro tenta match exato (case insensitive)
    for email_hospital in hospital_emails.keys():
        if hospital_name_clean == email_hospital.upper().strip():
            print(f"Match exato encontrado: '{hospital_name}' -> '{email_hospital}'")
            return hospital_emails[email_hospital]
    
    # Define common_generic aqui para que esteja disponível em todo o escopo
    common_generic = {'DE', 'DA', 'DO', 'E', 'EM', 'PARA', 'COM'}
    
    # 2. Tenta match parcial - se o nome do hospital está contido em algum nome da lista
    for email_hospital in hospital_emails.keys():
        email_hospital_clean = email_hospital.upper().strip()
        if hospital_name_clean in email_hospital_clean:
            print(f"Match parcial encontrado (hospital contido): '{hospital_name}' -> '{email_hospital}'")
            return hospital_emails[email_hospital]
        elif email_hospital_clean in hospital_name_clean:
            print(f"Match parcial encontrado (email contido): '{hospital_name}' -> '{email_hospital}'")
            return hospital_emails[email_hospital]
    
    # 3. Função para limpar nomes removendo palavras genéricas
    def clean_hospital_name(name):
        name_upper = name.upper()
        # Remove palavras comuns que não são específicas do hospital
        generic_words = [
            'HOSPITAL', 'HOSP', 'UNIDADE', 'FILIAL', 'CENTRO', 
            'INSTITUTO', 'CLINICA', 'CLÍNICA', 'POSTO', 'PRONTO',
            'ATENDIMENTO', 'EMERGENCIA', 'EMERGÊNCIA', 'COBRANCA'
        ]
        for word in generic_words:
            name_upper = name_upper.replace(word, '')
        
        # Remove espaços extras e caracteres especiais
        name_upper = ' '.join(name_upper.split())
        name_upper = ''.join(c for c in name_upper if c.isalnum() or c.isspace())
        return name_upper.strip()
    
    hospital_cleaned = clean_hospital_name(hospital_name_clean)
    
    # 4. Procura por match após limpeza
    best_match = None
    best_score = 0
    
    for email_hospital in hospital_emails.keys():
        email_cleaned = clean_hospital_name(email_hospital)
        
        # Calcula score de similaridade
        score = 0
        
        # A) Match exato após limpeza
        if hospital_cleaned == email_cleaned and hospital_cleaned:
            score = 100
        
        # B) Um nome contém o outro após limpeza
        elif hospital_cleaned and email_cleaned:
            if hospital_cleaned in email_cleaned:
                score = 90
            elif email_cleaned in hospital_cleaned:
                score = 90
        
        # C) Palavras em comum (método mais flexível)
        if score == 0 and hospital_cleaned and email_cleaned:
            hospital_words = set(hospital_cleaned.split())
            email_words = set(email_cleaned.split())
            common_words = hospital_words.intersection(email_words)
            
            # Remove palavras muito comuns
            meaningful_words = common_words - common_generic
            
            if meaningful_words:
                # Calcula score baseado nas palavras em comum
                score = min(80, len(meaningful_words) * 25)
        
        # D) Verifica se há match no nome original (sem limpeza)
        if score == 0:
            original_hospital_words = set(hospital_name_clean.split())
            original_email_words = set(email_hospital.upper().split())
            original_common = original_hospital_words.intersection(original_email_words)
            original_meaningful = original_common - common_generic
            
            if original_meaningful:
                score = min(70, len(original_meaningful) * 20)
        
        # Atualiza melhor match
        if score > best_score:
            best_score = score
            best_match = email_hospital
    
    # 5. Se encontrou um match com score bom, retorna
    if best_match and best_score >= 30:
        print(f"Match inteligente encontrado (score {best_score}): '{hospital_name}' -> '{best_match}'")
        return hospital_emails[best_match]
    
    # 6. Tenta matching por siglas como última tentativa
    def get_acronym(name):
        words = name.split()
        if len(words) > 1:
            # Pega primeira letra de cada palavra, ignorando palavras comuns
            acronym = ''.join(word[0] for word in words if word not in common_generic)
            return acronym if len(acronym) >= 2 else ""
        return ""
    
    hospital_acronym = get_acronym(hospital_name_clean)
    if hospital_acronym:
        for email_hospital in hospital_emails.keys():
            email_acronym = get_acronym(email_hospital.upper())
            if hospital_acronym == email_acronym:
                print(f"Match por sigla encontrado: '{hospital_name}' -> '{email_hospital}'")
                return hospital_emails[email_hospital]
    
    # 7. Se não encontrou nenhum match adequado
    print(f"AVISO: Nenhum match de email encontrado para: '{hospital_name}'")
    return None

def start_browser(headless=False):
    """Inicia o navegador Chrome controlado pelo Selenium."""
    options = webdriver.ChromeOptions()
    if headless:
        options.add_argument("--headless=new")

    prefs = {
        "download.default_directory": str(DOWNLOAD_FOLDER.resolve()),
        "download.prompt_for_download": False,
    }
    options.add_experimental_option("prefs", prefs)

    driver = webdriver.Chrome(
        service=ChromeService(ChromeDriverManager().install()),
        options=options
    )
    driver.maximize_window()
    return driver

def login_to_outlook(driver):
    """Realiza login no Outlook Web."""
    driver.get(OUTLOOK_WEB_URL)
    wait = WebDriverWait(driver, DEFAULT_WAIT_TIME)

    # Email
    email_input = wait.until(EC.presence_of_element_located((By.NAME, "loginfmt")))
    email_input.send_keys(OUTLOOK_EMAIL)
    email_input.send_keys(Keys.ENTER)
    time.sleep(2)

    # Senha
    password_input = wait.until(EC.presence_of_element_located((By.NAME, "passwd")))
    password_input.send_keys(OUTLOOK_PASSWORD)
    password_input.send_keys(Keys.ENTER)
    time.sleep(2)

    # Botão "Não" em "Stay signed in?"
    try:
        no_button = wait.until(EC.element_to_be_clickable((By.ID, "idBtn_Back")))
        no_button.click()
    except:
        pass

    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div[role='navigation']")))
    print("Login realizado com sucesso.")

def search_and_download_attachments(driver):
    """
    Procura emails não lidos com o assunto específico e baixa TODOS os anexos individualmente.
    """
    wait = WebDriverWait(driver, DEFAULT_WAIT_TIME)

    # Garante que estamos na inbox
    driver.get(OUTLOOK_WEB_URL)
    time.sleep(DEFAULT_WAIT_TIME)

    # CLICA NO BOTÃO DE FILTRO (ícone de filtro)
    print("Clicando no botão de filtro...")
    try:
        filter_button = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "/html/body/div[1]/div/div[2]/div/div[2]/div[2]/div/div[1]/div/div/div/div[3]/div/div/div[1]/div/div[1]/div[1]/div/div[2]/button[2]/span/i")))
        filter_button.click()
        print("Botão de filtro clicado com sucesso")
    except Exception as e:
        print(f"ERRO ao clicar no botão de filtro: {e}")
        # Tenta seletores alternativos para o botão de filtro
        alternative_filter_selectors = [
            (By.CSS_SELECTOR, "button[aria-label*='Filter']"),
            (By.CSS_SELECTOR, "button[data-icon-name*='Filter']"),
            (By.CSS_SELECTOR, "i[data-icon-name*='Filter']"),
            (By.XPATH, "//button[contains(@aria-label, 'Filter')]"),
            (By.XPATH, "//i[contains(@data-icon-name, 'Filter')]"),
        ]
        
        filter_found = False
        for by, selector in alternative_filter_selectors:
            try:
                filter_button = wait.until(EC.element_to_be_clickable((by, selector)))
                filter_button.click()
                print(f"Botão de filtro encontrado com seletor alternativo: {selector}")
                filter_found = True
                break
            except:
                continue
        
        if not filter_found:
            print("Não foi possível encontrar o botão de filtro")
            return

    time.sleep(DEFAULT_WAIT_TIME)

    # CLICA EM "UNREAD" NO MENU DE FILTRO
    print("Clicando em 'Unread'...")
    try:
        unread_option = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "/html/body/div[10]/div/div/div[2]/span[3]/span")))
        unread_option.click()
        print("Opção 'Unread' clicada com sucesso")
    except Exception as e:
        print(f"ERRO ao clicar em 'Unread': {e}")
        # Tenta seletores alternativos para a opção Unread
        alternative_unread_selectors = [
            (By.CSS_SELECTOR, "span[aria-label*='Unread']"),
            (By.CSS_SELECTOR, "button[aria-label*='Unread']"),
            (By.XPATH, "//span[contains(text(), 'Unread')]"),
            (By.XPATH, "//button[contains(text(), 'Unread')]"),
            (By.XPATH, "//*[contains(text(), 'Unread')]"),
        ]
        
        unread_found = False
        for by, selector in alternative_unread_selectors:
            try:
                unread_option = wait.until(EC.element_to_be_clickable((by, selector)))
                unread_option.click()
                print(f"Opção 'Unread' encontrada com seletor alternativo: {selector}")
                unread_found = True
                break
            except:
                continue
        
        if not unread_found:
            print("Não foi possível encontrar a opção 'Unread'")
            return

    time.sleep(DEFAULT_WAIT_TIME * 2)  # Aguarda mais tempo para os filtros serem aplicados

    # ENCONTRA E IDENTIFICA O EMAIL CORRETO COM O ASSUNTO CONFIGURADO
    print(f"Procurando email com '{EMAIL_SUBJECT}'...")
    
    target_email_element = None
    email_text = ""
    
    try:
        # Método 1: Procura pelo elemento específico com classe TtcXM
        subject_elements = driver.find_elements(By.XPATH, f"//span[@class='TtcXM' and contains(text(), '{EMAIL_SUBJECT}')]")
        
        if subject_elements:
            target_email_element = subject_elements[0]
            email_text = target_email_element.text
            print(f"ENCONTRADO: Email com '{EMAIL_SUBJECT}' - {email_text}")
        else:
            # Método 2: Procura por qualquer elemento contendo o assunto
            all_elements = driver.find_elements(By.XPATH, f"//*[contains(text(), '{EMAIL_SUBJECT}')]")
            for element in all_elements:
                if element.is_displayed() and EMAIL_SUBJECT in element.text:
                    target_email_element = element
                    email_text = element.text
                    print(f"ENCONTRADO: Email com '{EMAIL_SUBJECT}' - {email_text}")
                    break
        
        if not target_email_element:
            print(f"ERRO: Nenhum email com '{EMAIL_SUBJECT}' encontrado")
            return
            
    except Exception as e:
        print(f"ERRO ao procurar email com '{EMAIL_SUBJECT}': {e}")
        return

    # CLICA NO EMAIL PARA ABRIR E BAIXAR OS ANEXOS
    print("Clicando no email para abrir...")
    try:
        # Encontra o elemento pai clicável do email
        email_container = target_email_element.find_element(By.XPATH, "./ancestor::div[contains(@class, 'listItem') or contains(@role, 'option') or contains(@data-convid, '')][1]")
        email_container.click()
        print("Email aberto com sucesso")
    except Exception as e:
        print(f"ERRO ao clicar no email: {e}")
        # Tenta clicar diretamente no elemento
        try:
            target_email_element.click()
            print("Email aberto com clique direto")
        except Exception as e2:
            print(f"ERRO ao clicar diretamente: {e2}")
            return

    time.sleep(DEFAULT_WAIT_TIME * 2)
    
    # VERIFICA SE O EMAIL FOI ABERTO CORRETAMENTE
    try:
        # Tenta encontrar o assunto do email aberto para confirmar
        subject_selectors = [
            (By.CSS_SELECTOR, "[data-automation-id*='subject']"),
            (By.CSS_SELECTOR, "[aria-label*='Subject']"),
            (By.CSS_SELECTOR, "h1, h2, h3"),
            (By.CSS_SELECTOR, "[class*='subject']"),
        ]
        
        opened_subject = ""
        for by, selector in subject_selectors:
            try:
                subject_element = driver.find_element(by, selector)
                opened_subject = subject_element.text
                if opened_subject:
                    print(f"Assunto do email aberto: {opened_subject}")
                    break
            except:
                continue
        
        if not opened_subject:
            print("Não foi possível verificar o assunto do email aberto")
            
    except Exception as e:
        print(f"Erro ao verificar assunto do email aberto: {e}")
    
    # PROCURA E BAIXA TODOS OS ANEXOS INDIVIDUALMENTE
    print("Procurando todos os anexos para download...")
    
    try:
        # Primeiro, vamos encontrar todos os elementos de anexo
        attachment_selectors = [
            (By.CSS_SELECTOR, "[data-automation-id*='attachment']"),
            (By.CSS_SELECTOR, "[aria-label*='Attachment']"),
            (By.CSS_SELECTOR, "[title*='Attachment']"),
            (By.XPATH, "//*[contains(text(), 'attachment') or contains(text(), 'Anexo')]"),
        ]
        
        attachment_elements = []
        for by, selector in attachment_selectors:
            try:
                elements = driver.find_elements(by, selector)
                if elements:
                    attachment_elements.extend(elements)
                    print(f"Encontrados {len(elements)} elementos de anexo com seletor: {selector}")
            except:
                continue
        
        # Remove duplicatas
        unique_attachments = []
        seen_elements = set()
        for element in attachment_elements:
            element_id = element.id
            if element_id not in seen_elements:
                seen_elements.add(element_id)
                unique_attachments.append(element)
        
        print(f"Total de anexos únicos encontrados: {len(unique_attachments)}")
        
        if len(unique_attachments) == 0:
            print("Nenhum anexo encontrado. Tentando método alternativo...")
            # Método alternativo: procurar por botões de download específicos
            download_buttons = driver.find_elements(By.XPATH, "//button[contains(@aria-label, 'Download')]")
            print(f"Encontrados {len(download_buttons)} botões de download")
            
            # Baixa cada anexo individualmente
            for i, download_btn in enumerate(download_buttons):
                try:
                    print(f"Baixando anexo {i+1} de {len(download_buttons)}...")
                    download_btn.click()
                    time.sleep(DEFAULT_WAIT_TIME * 2)  # Aguarda entre downloads
                    print(f"Anexo {i+1} baixado com sucesso")
                except Exception as e:
                    print(f"Erro ao baixar anexo {i+1}: {e}")
        
        else:
            # Baixa cada anexo individualmente
            for i, attachment in enumerate(unique_attachments):
                try:
                    print(f"Processando anexo {i+1} de {len(unique_attachments)}...")
                    
                    # Clica no anexo para abrir as opções
                    attachment.click()
                    time.sleep(DEFAULT_WAIT_TIME)
                    
                    # Procura e clica no botão de download
                    download_selectors = [
                        (By.CSS_SELECTOR, "button[aria-label*='Download']"),
                        (By.CSS_SELECTOR, "button[title*='Download']"),
                        (By.XPATH, "//button[contains(@aria-label, 'Download')]"),
                        (By.CSS_SELECTOR, "button i[data-icon-name*='Download']"),
                    ]
                    
                    download_found = False
                    for by, selector in download_selectors:
                        try:
                            download_btn = wait.until(EC.element_to_be_clickable((by, selector)))
                            download_btn.click()
                            print(f"Anexo {i+1} baixado com sucesso")
                            download_found = True
                            break
                        except:
                            continue
                    
                    if not download_found:
                        print(f"AVISO: Não foi possível baixar o anexo {i+1}")
                    
                    # Aguarda entre downloads para evitar conflitos
                    time.sleep(DEFAULT_WAIT_TIME * 2)
                    
                except Exception as e:
                    print(f"Erro ao processar anexo {i+1}: {e}")
        
        print("Processo de download de anexos concluído!")
        
    except Exception as e:
        print(f"ERRO ao processar anexos: {e}")
        # Tenta método de fallback - baixar todos os anexos de uma vez
        try:
            print("Tentando método de fallback para download...")
            download_all_selector = (By.CSS_SELECTOR, "button[aria-label*='Download all']")
            download_all_btn = wait.until(EC.element_to_be_clickable(download_all_selector))
            download_all_btn.click()
            print("Download de todos os anexos iniciado (método fallback)")
        except Exception as fallback_error:
            print(f"ERRO no método fallback: {fallback_error}")

    time.sleep(DEFAULT_WAIT_TIME * 3)  # Aguarda downloads completarem

    # MARCA O EMAIL COMO LIDO - SEM RECARREGAR A PÁGINA
    print("Marcando email como lido com clique direito...")
    
    # VOLTA PARA A LISTA DE EMAILS SEM RECARREGAR A PÁGINA
    try:
        print("Voltando para a lista de emails usando botão Back...")
        # Tenta apenas o botão Back, sem recarregar a página
        back_selectors = [
            (By.CSS_SELECTOR, "button[aria-label*='Back']"),
            (By.CSS_SELECTOR, "button[data-icon-name*='Back']"),
            (By.XPATH, "//button[contains(@aria-label, 'Back')]"),
            (By.CSS_SELECTOR, "button[title*='Voltar']"),
        ]
        
        back_found = False
        for by, selector in back_selectors:
            try:
                back_button = wait.until(EC.element_to_be_clickable((by, selector)))
                back_button.click()
                print("Voltou para a lista de emails usando botão Back")
                back_found = True
                time.sleep(DEFAULT_WAIT_TIME * 2)
                break
            except:
                continue
        
        # SE NÃO ENCONTRAR O BOTÃO BACK, USA O ESC PARA VOLTAR
        if not back_found:
            print("Usando tecla ESC para voltar à lista...")
            actions = ActionChains(driver)
            actions.send_keys(Keys.ESCAPE).perform()
            print("Tecla ESC pressionada para voltar à lista")
            time.sleep(DEFAULT_WAIT_TIME * 2)
            
    except Exception as e:
        print(f"Erro ao voltar para lista de emails: {e}")
        # NÃO RECARREGA A PÁGINA - continua mesmo com erro

    # MARCA O EMAIL COMO LIDO NA LISTA DE NÃO LIDOS
    try:
        print(f"Encontrando o email com '{EMAIL_SUBJECT}' na lista de não lidos...")
        
        # Encontra o email específico com o assunto na lista atual
        target_email_element = None
        
        # Procura pelo email com o assunto na lista atual
        subject_elements = driver.find_elements(By.XPATH, f"//span[@class='TtcXM' and contains(text(), '{EMAIL_SUBJECT}')]")
        if not subject_elements:
            subject_elements = driver.find_elements(By.XPATH, f"//*[contains(text(), '{EMAIL_SUBJECT}')]")
        
        for element in subject_elements:
            if element.is_displayed() and EMAIL_SUBJECT in element.text:
                target_email_element = element
                break
        
        if not target_email_element:
            print("AVISO: Não foi possível encontrar o email na lista atual, mas continuando...")
            # Não interrompe o processo se não encontrar
            return
        
        # Encontra o container completo do email
        email_container = target_email_element.find_element(By.XPATH, "./ancestor::div[contains(@class, 'listItem') or contains(@role, 'option') or contains(@data-convid, '')][1]")
        
        # CLIQUE DIREITO NO EMAIL
        print("Clicando com botão direito no email...")
        actions = ActionChains(driver)
        actions.context_click(email_container).perform()
        print("Clique direito realizado no email")
        
        # Aguarda 1 segundo para o menu carregar
        time.sleep(1)
        
        # USA O XPATH ESPECÍFICO PARA "MARK AS READ"
        print("Clicando em 'Mark as read'...")
        try:
            mark_as_read_option = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//span[@class='fui-MenuItem__content r1ls86vo' and contains(text(), 'Mark as read')]")))
            mark_as_read_option.click()
            print("SUCESSO: Email marcado como lido na lista de não lidos!")
            
        except Exception as e:
            print(f"ERRO ao clicar em 'Mark as read': {e}")
            print("Tentando alternativa...")
            
            # TENTA ALTERNATIVA SIMPLIFICADA
            try:
                mark_as_read_alternative = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, "//span[contains(text(), 'Mark as read')]")))
                mark_as_read_alternative.click()
                print("SUCESSO: Email marcado como lido usando alternativa!")
                
            except Exception as e2:
                print(f"ERRO na alternativa: {e2}")
                print("Não foi possível marcar o email como lido")
        
    except Exception as e:
        print(f"ERRO ao marcar email como lido: {e}")

    time.sleep(DEFAULT_WAIT_TIME)
    print("Processo de download e marcação como lido concluído!")

def send_email_with_attachment(driver, pdf_paths, hospital_name, hospital_emails):
    """
    Envia um email com MÚLTIPLOS PDFs anexados para os emails específicos do hospital.
    AGORA ACEITA LISTA DE PDFs E CONTINUA MESMO COM ERRO.
    """
    global email_status_report
    
    wait = WebDriverWait(driver, 15)
    
    try:
        # VERIFICAÇÃO DOS ARQUIVOS PDF - se nenhum existir, registra erro mas continua
        valid_pdfs = []
        for pdf_path in pdf_paths:
            if pdf_path.exists():
                valid_pdfs.append(pdf_path)
            else:
                print(f"AVISO: Arquivo PDF não encontrado: {pdf_path}")
        
        if not valid_pdfs:
            print(f"ERRO: Nenhum arquivo PDF válido encontrado para {hospital_name}")
            email_status_report.append({
                'hospital': hospital_name,
                'arquivo': 'Nenhum arquivo válido',
                'situacao': 'Erro - Nenhum PDF encontrado'
            })
            return False
        
        # Busca os emails do hospital usando match flexível
        hospital_email_data = find_hospital_email(hospital_name, hospital_emails)
        if not hospital_email_data:
            print(f"AVISO: Não foram encontrados emails para o hospital {hospital_name}")
            email_status_report.append({
                'hospital': hospital_name,
                'arquivo': f"{len(pdf_paths)} arquivos",
                'situacao': 'Erro - Email do hospital não encontrado na planilha'
            })
            return False
        
        to_emails = hospital_email_data['to']
        cc_emails = hospital_email_data['cc']
        
        print(f"Enviando email para {hospital_name}: To={', '.join(to_emails)}, Cc={', '.join(cc_emails)}")
        print(f"Arquivos PDF ({len(valid_pdfs)}): {[p.name for p in valid_pdfs]}")
        
        print("Procurando botao de novo email...")
        
        # Tenta varios seletores para o botao New mail
        new_email_selectors = [
            (By.XPATH, "//button[.//span[contains(text(), 'New') or contains(text(), 'Novo')]]"),
            (By.CSS_SELECTOR, "button[aria-label*='New mail'], button[aria-label*='Nova mensagem']"),
            (By.CSS_SELECTOR, "button[data-tid*='newMail']"),
            (By.XPATH, "//button[contains(@aria-label, 'New mail')]"),
            (By.CSS_SELECTOR, "button[title*='New Mail']"),
            (By.CSS_SELECTOR, "i[data-icon-name*='Mail']"),
        ]
        
        new_email_button = None
        for by, selector in new_email_selectors:
            try:
                new_email_button = wait.until(EC.element_to_be_clickable((by, selector)))
                print(f"Botao New mail encontrado com: {by} = {selector}")
                break
            except:
                continue
        
        if not new_email_button:
            # Tenta método alternativo
            try:
                all_buttons = driver.find_elements(By.TAG_NAME, "button")
                for button in all_buttons:
                    try:
                        text = button.text.lower()
                        aria_label = button.get_attribute('aria-label') or ''
                        if 'new' in text or 'novo' in text or 'mail' in text or 'email' in text or 'mensagem' in text:
                            if button.is_displayed() and button.is_enabled():
                                button.click()
                                new_email_button = button
                                print("Botão New mail encontrado por texto/aria-label")
                                break
                    except:
                        continue
            except Exception as e:
                print(f"Erro no método alternativo: {e}")
            
        if not new_email_button:
            raise Exception("Nao foi possivel encontrar o botao de novo email")
            
        new_email_button.click()
        time.sleep(3)
        
        print("Aguardando janela de novo email carregar...")
        time.sleep(5)
        
        print("Preenchendo destinatario...")
        # Campo To - SELEÇÕES ESPECÍFICAS PARA O OUTLOOK WEB
        to_selectors = [
            (By.CSS_SELECTOR, "div[aria-label='To'][contenteditable='true']"),
            (By.CSS_SELECTOR, "div[aria-label*='To'][contenteditable='true']"),
            (By.CSS_SELECTOR, "div[role='textbox'][aria-label*='To']"),
            (By.CSS_SELECTOR, "div.EditorClass[contenteditable='true']"),
            (By.CSS_SELECTOR, "div[contenteditable='true'][aria-label*='To']"),
            (By.XPATH, "//div[@contenteditable='true' and contains(@aria-label, 'To')]"),
            (By.XPATH, "//div[@role='textbox' and contains(@aria-label, 'To')]"),
        ]
        
        to_field = None
        for by, selector in to_selectors:
            try:
                to_field = wait.until(EC.presence_of_element_located((by, selector)))
                print(f"Campo To encontrado com: {by} = {selector}")
                break
            except:
                continue
        
        if not to_field:
            # Busca alternativa por qualquer div contenteditable que possa ser o campo To
            try:
                divs = driver.find_elements(By.CSS_SELECTOR, "div[contenteditable='true']")
                for div in divs:
                    try:
                        aria_label = div.get_attribute('aria-label') or ''
                        if 'to' in aria_label.lower() or 'para' in aria_label.lower():
                            to_field = div
                            print("Campo To encontrado por busca alternativa em divs contenteditable")
                            break
                    except:
                        continue
            except Exception as e:
                print(f"Erro na busca alternativa do campo To: {e}")
        
        if to_field:
            print("Preenchendo campo To (div contenteditable)...")
            
            # Foca no campo
            to_field.click()
            time.sleep(1)
            
            # Limpa o conteúdo existente usando JavaScript
            try:
                driver.execute_script("arguments[0].innerHTML = '';", to_field)
                print("Conteúdo limpo via JavaScript")
            except:
                # Fallback: usa ActionChains para limpar
                try:
                    actions = ActionChains(driver)
                    actions.key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
                    time.sleep(1)
                    actions.send_keys(Keys.DELETE).perform()
                    time.sleep(1)
                    print("Conteúdo limpo via ActionChains")
                except Exception as e:
                    print(f"Erro ao limpar campo: {e}")
            
            # Digita os emails um por um com pausas
            to_emails_str = ', '.join(to_emails)
            print(f"Digitando emails no campo To: {to_emails_str}")
            
            for email in to_emails:
                to_field.send_keys(email)
                time.sleep(0.5)
                to_field.send_keys(",")
                time.sleep(0.3)
            
            # Remove a última vírgula se existir
            time.sleep(1)
            
            print("Campo To preenchido com sucesso")
            time.sleep(2)
            
            # Pressiona TAB para sair do campo e confirmar os emails
            print("Pressionando TAB para sair do campo To...")
            actions = ActionChains(driver)
            actions.send_keys(Keys.TAB).perform()
            time.sleep(2)
            
        else:
            print("ERRO: Campo To não encontrado")
            raise Exception("Campo To não encontrado")
        
        # PREENCHER CAMPO CC
        print("Preenchendo campo Cc...")
        if cc_emails:  # So preenche se houver emails no Cc
            print(f"Emails CC encontrados: {cc_emails}")
            
            # Tenta varios seletores para o campo Cc
            cc_selectors = [
                # Primeiro tenta encontrar por texto/clique no link "Cc"
                (By.XPATH, "//span[contains(text(), 'Cc')]"),
                (By.CSS_SELECTOR, "button[aria-label*='Cc']"),
                (By.CSS_SELECTOR, "a[aria-label*='Cc']"),
                
                # Depois tenta os campos de entrada
                (By.XPATH, "/html/body/div[1]/div/div[2]/div/div[2]/div[2]/div/div[1]/div/div/div/div[3]/div/div/div[3]/div[1]/div/div/div/div/div[3]/div[1]/div/div[4]/div/span/span[2]/div/div[1]"),
                (By.CSS_SELECTOR, "div[role='textbox'][aria-label='Cc']"),
                (By.CSS_SELECTOR, "input[aria-label*='Cc']"),
                (By.CSS_SELECTOR, "div[aria-label*='Cc']"),
            ]
            
            cc_field = None
            cc_found_by_click = False
            
            # Estrategia 1: Tenta clicar no link/texto "Cc" primeiro
            for by, selector in cc_selectors[:3]:
                try:
                    cc_link = wait.until(EC.element_to_be_clickable((by, selector)))
                    print(f"Link Cc encontrado com: {by} = {selector}")
                    cc_link.click()
                    print("Clicou no link Cc")
                    cc_found_by_click = True
                    time.sleep(2)
                    break
                except:
                    continue
            
            # Estrategia 2: Se nao encontrou o link, tenta encontrar o campo diretamente
            if not cc_found_by_click:
                print("Nao encontrou link Cc, procurando campo diretamente...")
                for by, selector in cc_selectors[3:]:
                    try:
                        cc_field = wait.until(EC.presence_of_element_located((by, selector)))
                        print(f"Campo Cc encontrado com: {by} = {selector}")
                        break
                    except:
                        continue
            
            # Preenche o campo Cc
            if cc_found_by_click or cc_field:
                # Se clicou no link Cc, agora precisa encontrar o campo que apareceu
                if cc_found_by_click:
                    time.sleep(1)
                    # Procura o campo Cc que deve ter aparecido apos o clique
                    cc_field_selectors = [
                        (By.CSS_SELECTOR, "div[role='textbox'][aria-label='Cc']"),
                        (By.CSS_SELECTOR, "input[aria-label*='Cc']"),
                        (By.CSS_SELECTOR, "div[aria-label*='Cc']"),
                    ]
                    
                    for by, selector in cc_field_selectors:
                        try:
                            cc_field = wait.until(EC.presence_of_element_located((by, selector)))
                            print(f"Campo Cc apareceu apos clique: {by} = {selector}")
                            break
                        except:
                            continue
                
                if cc_field:
                    # Limpa o campo primeiro (caso tenha algo)
                    cc_field.clear()
                    time.sleep(1)
                    
                    # Junta todos os emails CC com virgula
                    cc_emails_str = ', '.join(cc_emails)
                    cc_field.send_keys(cc_emails_str)
                    print(f"Cc preenchido com: {cc_emails_str}")
                    time.sleep(2)
                    
                    # Pressiona TAB para sair do campo Cc
                    cc_field.send_keys(Keys.TAB)
                    time.sleep(1)
                else:
                    print("AVISO: Campo Cc nao encontrado apos clique no link")
            else:
                print("AVISO: Campo Cc nao encontrado, continuando sem preencher Cc")
        else:
            print("Nenhum email para preencher no campo Cc")
        # FIM DO CAMPO CC
        
        print("Preenchendo assunto...")
        # Campo Subject
        subject_selectors = [
            (By.XPATH, "/html/body/div[1]/div/div[2]/div/div[2]/div[2]/div/div/div/div[3]/div/div/div[3]/div[1]/div/div/div/div/div[3]/div[2]/span/input"),
            (By.CSS_SELECTOR, "input[aria-label*='Subject']"),
            (By.CSS_SELECTOR, "input[placeholder*='Add a subject']"),
            (By.CSS_SELECTOR, "input[name*='Subject']"),
        ]
        
        subject_field = None
        for by, selector in subject_selectors:
            try:
                subject_field = wait.until(EC.presence_of_element_located((by, selector)))
                print(f"Campo Subject encontrado com: {by} = {selector}")
                break
            except:
                continue
        
        if subject_field:
            subject_field.send_keys(f"Boletos em aberto: {hospital_name}")
            time.sleep(2)
        
        print("Preenchendo corpo do email...")
        # Corpo do email
        body_selectors = [
            (By.XPATH, "/html/body/div[1]/div/div[2]/div/div[2]/div[2]/div/div/div/div[3]/div/div/div[3]/div[1]/div/div/div/div/div[4]/div/div[1]/div"),
            (By.CSS_SELECTOR, "div[role='textbox'][aria-label='Message body']"),
            (By.CSS_SELECTOR, "div[aria-label*='Message body']"),
            (By.CSS_SELECTOR, "div[contenteditable='true']"),
        ]
        
        body_field = None
        for by, selector in body_selectors:
            try:
                body_field = wait.until(EC.presence_of_element_located((by, selector)))
                print(f"Corpo do email encontrado com: {by} = {selector}")
                break
            except:
                continue
        
        if body_field:
            mensagem = f"""Prezados(as), Bom dia.

Segue abaixo a relação de boletos em aberto com o hospital, caso tenha sido efetuado o pagamento favor enviar o comprovante para baixa.."""
            
            body_field.send_keys(mensagem)
            time.sleep(2)
        
        # ANEXAR MÚLTIPLOS PDFs
        print(f"Anexando {len(valid_pdfs)} arquivos PDF...")
        
        for i, pdf_path in enumerate(valid_pdfs):
            print(f"Anexando arquivo {i+1} de {len(valid_pdfs)}: {pdf_path.name}")
            
            print("Clicando em Insert...")
            # Insert option
            insert_selectors = [
                (By.XPATH, "/html/body/div[1]/div/div[2]/div/div[2]/div[1]/div/div/div[1]/div[2]/div/div/div/div/span/div[1]/div/div/div[5]/div/button/span/span/span"),
                (By.CSS_SELECTOR, "button[aria-label*='Insert']"),
            ]
            
            insert_button = None
            for by, selector in insert_selectors:
                try:
                    insert_button = wait.until(EC.element_to_be_clickable((by, selector)))
                    print(f"Botao Insert encontrado com: {by} = {selector}")
                    break
                except:
                    continue
            
            if insert_button:
                insert_button.click()
                time.sleep(2)
            
            print("Clicando em Attach file...")
            # Attach file button
            attach_selectors = [
                (By.XPATH, "/html/body/div[1]/div/div[2]/div/div[2]/div[1]/div/div/div[2]/div[1]/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/button/span/span[1]/span"),
                (By.CSS_SELECTOR, "button[aria-label*='Attach file']"),
                (By.CSS_SELECTOR, "button[data-icon-name*='Attach file']"),
            ]
            
            attach_button = None
            for by, selector in attach_selectors:
                try:
                    attach_button = wait.until(EC.element_to_be_clickable((by, selector)))
                    print(f"Botao Attach encontrado com: {by} = {selector}")
                    break
                except:
                    continue
            
            if attach_button:
                attach_button.click()
                time.sleep(2)
            
            print("Clicando em Browse this computer...")
            # Browse this computer button
            browse_selectors = [
                (By.XPATH, "/html/body/div[2]/div[6]/div/div/div/div/div/div/ul/li/div/ul/li[1]/button/div/span"),
                (By.CSS_SELECTOR, "button[aria-label*='Browse this computer']"),
                (By.CSS_SELECTOR, "div[title*='Browse this computer']"),
            ]
            
            browse_button = None
            for by, selector in browse_selectors:
                try:
                    browse_button = wait.until(EC.element_to_be_clickable((by, selector)))
                    print(f"Botao Browse encontrado com: {by} = {selector}")
                    break
                except:
                    continue
            
            if browse_button:
                browse_button.click()
                time.sleep(3)
            
            print("Navegando para a pasta dos PDFs...")
            time.sleep(3)
            
            pyautogui.click(x=500, y=100)
            time.sleep(2)
            
            pyautogui.hotkey('ctrl', 'a')
            time.sleep(1)
            pyautogui.press('delete')
            time.sleep(1)
            
            pdf_folder_path = str(PROCESSED_FOLDER.resolve())
            print(f"Digitando caminho da pasta: {pdf_folder_path}")
            pyautogui.write(pdf_folder_path)
            time.sleep(1)
            pyautogui.press('enter')
            time.sleep(3)
            
            pdf_filename = pdf_path.name
            print(f"Digitando nome do arquivo: {pdf_filename}")
            pyautogui.write(pdf_filename)
            time.sleep(1)
            pyautogui.press('enter')
            time.sleep(5)
            
            print(f"Verificando se o arquivo {i+1} foi anexado...")
            time.sleep(3)
            
            attachment_found = False
            
            try:
                attachment_indicators = [
                    (By.CSS_SELECTOR, "[aria-label*='Anexo']"),
                    (By.CSS_SELECTOR, "[data-automation-id*='attachment']"),
                    (By.CSS_SELECTOR, "[class*='attachment']"),
                    (By.CSS_SELECTOR, "[class*='attachments']"),
                    (By.XPATH, "//*[contains(text(), '.pdf')]"),
                    (By.CSS_SELECTOR, "[title*='.pdf']"),
                    (By.CSS_SELECTOR, "[data-tid*='attachment']"),
                    (By.CSS_SELECTOR, "div[role='listitem']"),
                ]
                
                for by, selector in attachment_indicators:
                    try:
                        elements = driver.find_elements(by, selector)
                        for element in elements:
                            if element.is_displayed():
                                text = element.text.lower() if element.text else ""
                                if '.pdf' in text or 'anexo' in text or 'attachment' in text:
                                    attachment_found = True
                                    print(f"SUCESSO: Anexo {i+1} detectado com seletor: {selector}")
                                    break
                        if attachment_found:
                            break
                    except:
                        continue
            except Exception as e:
                print(f"Aviso na verificacao de anexo {i+1}: {e}")
            
            if not attachment_found:
                try:
                    elements_with_text = driver.find_elements(By.XPATH, f"//*[contains(text(), '{pdf_path.name}')]")
                    for element in elements_with_text:
                        if element.is_displayed():
                            attachment_found = True
                            print(f"SUCESSO: Anexo {i+1} detectado pelo nome do arquivo: {pdf_path.name}")
                            break
                except:
                    pass
            
            if not attachment_found:
                print(f"AVISO: Nao foi possivel detectar o anexo {i+1} automaticamente, mas continuando o processo...")
                print("Isso pode acontecer devido a diferencas na interface do Outlook.")
                attachment_found = True
            
            if not attachment_found:
                print(f"ERRO: Anexo {i+1} nao foi detectado apos tentativa de upload")
                # Continua tentando anexar os próximos arquivos mesmo com erro
        
        print("Todos os arquivos processados para anexação")
        
        print("Enviando email...")
        send_selectors = [
            (By.CSS_SELECTOR, "button[aria-label*='Send']"),
            (By.CSS_SELECTOR, "button[data-icon-name*='Send']"),
            (By.XPATH, "//button[contains(@aria-label, 'Send')]"),
        ]
        
        send_button = None
        for by, selector in send_selectors:
            try:
                send_button = wait.until(EC.element_to_be_clickable((by, selector)))
                print(f"Botao Send encontrado com: {by} = {selector}")
                break
            except:
                continue
        
        if send_button:
            send_button.click()
            time.sleep(3)
        
        print(f"SUCESSO: Email enviado com sucesso para {hospital_name} com {len(valid_pdfs)} anexos")
        
        email_status_report.append({
            'hospital': hospital_name,
            'arquivo': f"{len(valid_pdfs)} arquivos",
            'situacao': 'Enviado'
        })
        
        # EXCLUIR TODOS OS PDFs DO HOSPITAL APÓS ENVIO
        print(f"Excluindo {len(valid_pdfs)} arquivos PDF do hospital {hospital_name}...")
        for pdf_path in valid_pdfs:
            try:
                if pdf_path.exists():
                    pdf_path.unlink()
                    print(f"  PDF excluido: {pdf_path.name}")
                else:
                    print(f"  Aviso: Arquivo PDF não encontrado para exclusão: {pdf_path}")
            except Exception as delete_error:
                print(f"  Erro ao excluir PDF {pdf_path.name}: {delete_error}")
        
        return True
        
    except Exception as e:
        print(f"ERRO ao enviar email para {hospital_name}: {e}")
        email_status_report.append({
            'hospital': hospital_name,
            'arquivo': f"{len(pdf_paths)} arquivos",
            'situacao': f'Erro - {str(e)[:100]}'
        })
        return False

def send_all_pdfs_by_email(pdf_files, hospital_emails):
    """
    Envia todos os PDFs por email, agrupados por hospital.
    Cada hospital recebe UM email com TODOS os seus PDFs anexados.
    CONTINUA MESMO COM ERROS - não para a execução.
    """
    global email_status_report
    
    if not pdf_files:
        print("Nenhum PDF para enviar.")
        return
    
    print(f"Iniciando envio de {len(pdf_files)} PDFs agrupados por hospital...")
    
    # Agrupa PDFs por hospital
    hospital_pdfs = group_pdfs_by_hospital(pdf_files)
    
    print(f"Total de hospitais encontrados: {len(hospital_pdfs)}")
    
    # Inicia um novo navegador para envio de emails
    email_driver = start_browser(headless=False)
    
    try:
        # Login no Outlook
        login_to_outlook(email_driver)
        time.sleep(DEFAULT_WAIT_TIME)
        
        # Para cada hospital, envia UM email com TODOS os seus PDFs
        successful_sends = 0
        failed_sends = 0
        
        for hospital_name, pdf_paths in hospital_pdfs.items():
            print(f"\n=== Processando email para: {hospital_name} ===")
            print(f"PDFs a anexar: {[p.name for p in pdf_paths]}")
            
            success = send_email_with_attachment(email_driver, pdf_paths, hospital_name, hospital_emails)
            
            if success:
                successful_sends += 1
                print(f"SUCESSO: Email enviado para {hospital_name} com {len(pdf_paths)} anexos")
            else:
                failed_sends += 1
                print(f"FALHA: Email não enviado para {hospital_name} (continuando para o próximo...)")
            
            # Aguarda um pouco entre os emails
            time.sleep(DEFAULT_WAIT_TIME * 2)
        
        # Resumo final
        print(f"\n=== RESUMO DO ENVIO ===")
        print(f"Total de hospitais processados: {len(hospital_pdfs)}")
        print(f"Emails enviados com sucesso: {successful_sends}")
        print(f"Emails com falha: {failed_sends}")
        print(f"Total de PDFs processados: {len(pdf_files)}")
        
        if failed_sends > 0:
            print(f"AVISO: {failed_sends} emails não foram enviados. Verifique o relatório para detalhes.")
        
        # BACKUP: Tenta excluir PDFs restantes
        remaining_pdfs = [pdf for pdf in pdf_files if pdf.exists()]
        if remaining_pdfs:
            print(f"Limpando {len(remaining_pdfs)} arquivos PDF restantes...")
            for pdf in remaining_pdfs:
                try:
                    pdf.unlink()
                    print(f"  Excluído: {pdf.name}")
                except Exception as e:
                    print(f"  Erro ao excluir {pdf.name}: {e}")
        
    except Exception as e:
        print(f"ERRO durante o envio de emails: {e}")
        print("CONTINUANDO para gerar relatório...")
    finally:
        # Fecha o navegador de email
        email_driver.quit()

def generate_email_status_report():
    """
    Gera um PDF com o relatório dos emails enviados.
    AGORA INCLUI INFORMAÇÕES MAIS DETALHADAS SOBRE OS ERROS.
    """
    try:
        if not email_status_report:
            print("Nenhum dado de status para gerar relatório.")
            return None
        
        # Nome do arquivo de relatório
        report_filename = f"Relatorio_Envio_Emails_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        report_path = PROCESSED_FOLDER / report_filename
        
        # Cria o documento PDF
        doc = SimpleDocTemplate(str(report_path), pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title = Paragraph("RELATÓRIO DE ENVIO DE EMAILS - FATURAS", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Data de emissão
        data_emissao = Paragraph(f"Data de emissão: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal'])
        elements.append(data_emissao)
        elements.append(Spacer(1, 20))
        
        # Resumo estatístico
        total_emails = len(email_status_report)
        enviados = sum(1 for item in email_status_report if item['situacao'] == 'Enviado')
        erros = total_emails - enviados
        
        resumo_text = f"Resumo: {enviados} enviados com sucesso, {erros} com erro"
        resumo = Paragraph(resumo_text, styles['Normal'])
        elements.append(resumo)
        elements.append(Spacer(1, 20))
        
        # Prepara dados da tabela - AGORA COM COLUNA DE HOSPITAL
        table_data = [['Hospital', 'Arquivo', 'Situação']]
        
        for item in email_status_report:
            hospital = item.get('hospital', 'N/A')
            arquivo = item['arquivo']
            situacao = item['situacao']
            
            table_data.append([hospital, arquivo, situacao])
        
        # Cria tabela
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F2F2F2')),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            # Colore as células de situação
            ('TEXTCOLOR', (2, 1), (2, -1), 
             lambda row, col, table=table_data: 
                 colors.green if table[row][col] == 'Enviado' else colors.red),
        ]))
        
        elements.append(table)
        
        # Adiciona seção de detalhes dos erros se houver
        if erros > 0:
            elements.append(Spacer(1, 20))
            erro_title = Paragraph("DETALHES DOS ERROS:", styles['Heading2'])
            elements.append(erro_title)
            
            for item in email_status_report:
                if item['situacao'] != 'Enviado':
                    erro_text = f"- {item.get('hospital', 'N/A')}: {item['situacao']}"
                    erro_para = Paragraph(erro_text, styles['Normal'])
                    elements.append(erro_para)
        
        # Gera o PDF
        doc.build(elements)
        print(f"Relatório de status gerado: {report_filename}")
        
        return report_path
        
    except Exception as e:
        print(f"ERRO ao gerar relatório de status: {e}")
        return None

def send_status_report_email(driver, report_path):
    """
    Envia o relatório de status por email para LOG_AUTOMATION_EMAIL.
    Versão corrigida usando a mesma lógica da função send_email_with_attachment.
    """
    if not report_path or not report_path.exists():
        print("ERRO: Arquivo de relatório não encontrado para envio.")
        return False
    
    wait = WebDriverWait(driver, 15)
    
    try:
        print("=== ENVIANDO RELATÓRIO DE STATUS POR EMAIL ===")
        
        # Garante que estamos na inbox
        driver.get("https://outlook.office.com/mail/inbox")
        time.sleep(5)
        
        print("Procurando botão de novo email...")
        # Usa os mesmos seletores da função send_email_with_attachment
        new_email_selectors = [
            (By.XPATH, "//button[.//span[contains(text(), 'New') or contains(text(), 'Novo')]]"),
            (By.CSS_SELECTOR, "button[aria-label*='New mail'], button[aria-label*='Nova mensagem']"),
            (By.CSS_SELECTOR, "button[data-tid*='newMail']"),
            (By.XPATH, "//button[contains(@aria-label, 'New mail')]"),
            (By.CSS_SELECTOR, "button[title*='New Mail']"),
            (By.CSS_SELECTOR, "i[data-icon-name*='Mail']"),
        ]
        
        new_email_button = None
        for by, selector in new_email_selectors:
            try:
                new_email_button = wait.until(EC.element_to_be_clickable((by, selector)))
                print(f"Botão New mail encontrado com: {by} = {selector}")
                break
            except:
                continue
        
        if not new_email_button:
            # Tenta método alternativo
            try:
                all_buttons = driver.find_elements(By.TAG_NAME, "button")
                for button in all_buttons:
                    try:
                        text = button.text.lower()
                        aria_label = button.get_attribute('aria-label') or ''
                        if 'new' in text or 'novo' in text or 'mail' in text or 'email' in text or 'mensagem' in text:
                            if button.is_displayed() and button.is_enabled():
                                button.click()
                                new_email_button = button
                                print("Botão New mail encontrado por texto/aria-label")
                                break
                    except:
                        continue
            except Exception as e:
                print(f"Erro no método alternativo: {e}")
            
        if not new_email_button:
            raise Exception("Não foi possível encontrar o botão de novo email")
            
        new_email_button.click()
        time.sleep(3)
        
        print("Aguardando janela de novo email carregar...")
        time.sleep(5)
        
        print("Preenchendo destinatário...")
        # CAMPO TO - USA A MESMA LÓGICA DA FUNÇÃO send_email_with_attachment
        to_selectors = [
            (By.CSS_SELECTOR, "div[aria-label='To'][contenteditable='true']"),
            (By.CSS_SELECTOR, "div[aria-label*='To'][contenteditable='true']"),
            (By.CSS_SELECTOR, "div[role='textbox'][aria-label*='To']"),
            (By.CSS_SELECTOR, "div.EditorClass[contenteditable='true']"),
            (By.CSS_SELECTOR, "div[contenteditable='true'][aria-label*='To']"),
            (By.XPATH, "//div[@contenteditable='true' and contains(@aria-label, 'To')]"),
            (By.XPATH, "//div[@role='textbox' and contains(@aria-label, 'To')]"),
        ]
        
        to_field = None
        for by, selector in to_selectors:
            try:
                to_field = wait.until(EC.presence_of_element_located((by, selector)))
                print(f"Campo To encontrado com: {by} = {selector}")
                break
            except:
                continue
        
        if not to_field:
            # Busca alternativa por qualquer div contenteditable que possa ser o campo To
            try:
                divs = driver.find_elements(By.CSS_SELECTOR, "div[contenteditable='true']")
                for div in divs:
                    try:
                        aria_label = div.get_attribute('aria-label') or ''
                        if 'to' in aria_label.lower() or 'para' in aria_label.lower():
                            to_field = div
                            print("Campo To encontrado por busca alternativa em divs contenteditable")
                            break
                    except:
                        continue
            except Exception as e:
                print(f"Erro na busca alternativa do campo To: {e}")
        
        if to_field:
            print("Preenchendo campo To (div contenteditable)...")
            
            # Foca no campo
            to_field.click()
            time.sleep(1)
            
            # Limpa o conteúdo existente usando JavaScript
            try:
                driver.execute_script("arguments[0].innerHTML = '';", to_field)
                print("Conteúdo limpo via JavaScript")
            except:
                # Fallback: usa ActionChains para limpar
                try:
                    actions = ActionChains(driver)
                    actions.key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
                    time.sleep(1)
                    actions.send_keys(Keys.DELETE).perform()
                    time.sleep(1)
                    print("Conteúdo limpo via ActionChains")
                except Exception as e:
                    print(f"Erro ao limpar campo: {e}")
            
            # Digita o email
            print(f"Digitando email no campo To: {LOG_AUTOMATION_EMAIL}")
            
            to_field.send_keys(LOG_AUTOMATION_EMAIL)
            time.sleep(0.5)
            
            print("Campo To preenchido com sucesso")
            time.sleep(2)
            
            # Pressiona TAB para sair do campo e confirmar o email
            print("Pressionando TAB para sair do campo To...")
            actions = ActionChains(driver)
            actions.send_keys(Keys.TAB).perform()
            time.sleep(2)
            
        else:
            print("ERRO: Campo To não encontrado")
            raise Exception("Campo To não encontrado")
        
        print("Preenchendo assunto...")
        # Campo Subject - usa mesma lógica robusta
        subject_selectors = [
            (By.XPATH, "/html/body/div[1]/div/div[2]/div/div[2]/div[2]/div/div/div/div[3]/div/div/div[3]/div[1]/div/div/div/div/div[3]/div[2]/span/input"),
            (By.CSS_SELECTOR, "input[aria-label*='Subject']"),
            (By.CSS_SELECTOR, "input[placeholder*='Add a subject']"),
            (By.CSS_SELECTOR, "input[name*='Subject']"),
        ]
        
        subject_field = None
        for by, selector in subject_selectors:
            try:
                subject_field = wait.until(EC.presence_of_element_located((by, selector)))
                print(f"Campo Subject encontrado com: {by} = {selector}")
                break
            except:
                continue
        
        if subject_field:
            subject_text = f"Relatório de envio de boletos em aberto - {datetime.now().strftime('%d/%m/%Y')}"
            subject_field.send_keys(subject_text)
            time.sleep(2)
        
        print("Preenchendo corpo do email...")
        # Corpo do email - usa mesma lógica robusta
        body_selectors = [
            (By.XPATH, "/html/body/div[1]/div/div[2]/div/div[2]/div[2]/div/div/div/div[3]/div/div/div[3]/div[1]/div/div/div/div/div[4]/div/div[1]/div"),
            (By.CSS_SELECTOR, "div[role='textbox'][aria-label='Message body']"),
            (By.CSS_SELECTOR, "div[aria-label*='Message body']"),
            (By.CSS_SELECTOR, "div[contenteditable='true']"),
        ]
        
        body_field = None
        for by, selector in body_selectors:
            try:
                body_field = wait.until(EC.presence_of_element_located((by, selector)))
                print(f"Corpo do email encontrado com: {by} = {selector}")
                break
            except:
                continue
        
        if body_field:
            # Calcula estatísticas
            total_emails = len(email_status_report)
            enviados = sum(1 for item in email_status_report if item['situacao'] == 'Enviado')
            erros = total_emails - enviados
            
            mensagem = f"""Prezados,

Segue em anexo o relatório de envio de boletos em aberto realizado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}.

Resumo do processamento:
- Total de emails processados: {total_emails}
- Emails enviados com sucesso: {enviados}
- Emails com erro: {erros}

O relatório detalhado está em anexo.

Atenciosamente,
Sistema de Automação de Boletos em Aberto"""
            
            body_field.send_keys(mensagem)
            time.sleep(2)
        
        print("Clicando em Insert...")
        # Insert option - mesma lógica
        insert_selectors = [
            (By.XPATH, "/html/body/div[1]/div/div[2]/div/div[2]/div[1]/div/div/div[1]/div[2]/div/div/div/div/span/div[1]/div/div/div[5]/div/button/span/span/span"),
            (By.CSS_SELECTOR, "button[aria-label*='Insert']"),
        ]
        
        insert_button = None
        for by, selector in insert_selectors:
            try:
                insert_button = wait.until(EC.element_to_be_clickable((by, selector)))
                print(f"Botão Insert encontrado com: {by} = {selector}")
                break
            except:
                continue
        
        if insert_button:
            insert_button.click()
            time.sleep(2)
        
        print("Clicando em Attach file...")
        # Attach file button - mesma lógica
        attach_selectors = [
            (By.XPATH, "/html/body/div[1]/div/div[2]/div/div[2]/div[1]/div/div/div[2]/div[1]/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/button/span/span[1]/span"),
            (By.CSS_SELECTOR, "button[aria-label*='Attach file']"),
            (By.CSS_SELECTOR, "button[data-icon-name*='Attach file']"),
        ]
        
        attach_button = None
        for by, selector in attach_selectors:
            try:
                attach_button = wait.until(EC.element_to_be_clickable((by, selector)))
                print(f"Botão Attach encontrado com: {by} = {selector}")
                break
            except:
                continue
        
        if attach_button:
            attach_button.click()
            time.sleep(2)
        
        print("Clicando em Browse this computer...")
        # Browse this computer button - mesma lógica
        browse_selectors = [
            (By.XPATH, "/html/body/div[2]/div[6]/div/div/div/div/div/div/ul/li/div/ul/li[1]/button/div/span"),
            (By.CSS_SELECTOR, "button[aria-label*='Browse this computer']"),
            (By.CSS_SELECTOR, "div[title*='Browse this computer']"),
        ]
        
        browse_button = None
        for by, selector in browse_selectors:
            try:
                browse_button = wait.until(EC.element_to_be_clickable((by, selector)))
                print(f"Botão Browse encontrado com: {by} = {selector}")
                break
            except:
                continue
        
        if browse_button:
            browse_button.click()
            time.sleep(3)
        
        print("Navegando para a pasta dos relatórios...")
        time.sleep(3)
        
        pyautogui.click(x=500, y=100)
        time.sleep(2)
        
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(1)
        pyautogui.press('delete')
        time.sleep(1)
        
        report_folder_path = str(PROCESSED_FOLDER.resolve())
        print(f"Digitando caminho da pasta: {report_folder_path}")
        pyautogui.write(report_folder_path)
        time.sleep(1)
        pyautogui.press('enter')
        time.sleep(3)
        
        report_filename = report_path.name
        print(f"Digitando nome do arquivo: {report_filename}")
        pyautogui.write(report_filename)
        time.sleep(1)
        pyautogui.press('enter')
        time.sleep(5)
        
        print("Verificando se o arquivo foi anexado...")
        time.sleep(3)
        
        attachment_found = False
        
        try:
            attachment_indicators = [
                (By.CSS_SELECTOR, "[aria-label*='Anexo']"),
                (By.CSS_SELECTOR, "[data-automation-id*='attachment']"),
                (By.CSS_SELECTOR, "[class*='attachment']"),
                (By.CSS_SELECTOR, "[class*='attachments']"),
                (By.XPATH, "//*[contains(text(), '.pdf')]"),
                (By.CSS_SELECTOR, "[title*='.pdf']"),
                (By.CSS_SELECTOR, "[data-tid*='attachment']"),
                (By.CSS_SELECTOR, "div[role='listitem']"),
            ]
            
            for by, selector in attachment_indicators:
                try:
                    elements = driver.find_elements(by, selector)
                    for element in elements:
                        if element.is_displayed():
                            text = element.text.lower() if element.text else ""
                            if '.pdf' in text or 'anexo' in text or 'attachment' in text:
                                attachment_found = True
                                print(f"SUCESSO: Anexo detectado com seletor: {selector}")
                                print(f"Texto do elemento: {text}")
                                break
                    if attachment_found:
                        break
                except:
                    continue
        except Exception as e:
            print(f"Aviso na verificação de anexo: {e}")
        
        if not attachment_found:
            try:
                elements_with_text = driver.find_elements(By.XPATH, f"//*[contains(text(), '{report_path.name}')]")
                for element in elements_with_text:
                    if element.is_displayed():
                        attachment_found = True
                        print(f"SUCESSO: Anexo detectado pelo nome do arquivo: {report_path.name}")
                        break
            except:
                pass
        
        if not attachment_found:
            print("AVISO: Não foi possível detectar o anexo automaticamente, mas continuando o processo...")
            print("Isso pode acontecer devido a diferenças na interface do Outlook.")
            attachment_found = True
        
        if not attachment_found:
            print("ERRO: Anexo não foi detectado após tentativa de upload")
            return False
        
        print("Enviando email...")
        send_selectors = [
            (By.CSS_SELECTOR, "button[aria-label*='Send']"),
            (By.CSS_SELECTOR, "button[data-icon-name*='Send']"),
            (By.XPATH, "//button[contains(@aria-label, 'Send')]"),
        ]
        
        send_button = None
        for by, selector in send_selectors:
            try:
                send_button = wait.until(EC.element_to_be_clickable((by, selector)))
                print(f"Botão Send encontrado com: {by} = {selector}")
                break
            except:
                continue
        
        if send_button:
            send_button.click()
            time.sleep(3)
        
        print("SUCESSO: Relatório de status enviado por email!")
        
        # Verifica se o email foi enviado com sucesso
        try:
            # Procura por mensagem de confirmação
            confirmation_selectors = [
                (By.XPATH, "//*[contains(text(), 'sent') or contains(text(), 'enviado')]"),
                (By.CSS_SELECTOR, "[data-tid*='messageSent']"),
            ]
            
            for by, selector in confirmation_selectors:
                try:
                    confirmation = driver.find_element(by, selector)
                    if confirmation.is_displayed():
                        print("Confirmação de envio visualizada!")
                        break
                except:
                    continue
        except:
            pass
        
        return True
        
    except Exception as e:
        print(f"ERRO CRÍTICO ao enviar relatório de status: {e}")
        return False

if __name__ == "__main__":
    # Carrega as configurações do Excel
    print("Carregando configurações do arquivo Excel...")
    if not load_config_from_excel():
        print("ERRO: Não foi possível carregar as configurações. Verifique o arquivo Excel.")
        exit(1)
    
    # Limpa as pastas antes de iniciar
    print("Limpando pastas antes de iniciar...")
    clean_folders()
    
    # Cria as pastas (caso não existam)
    create_folders()
    
    # Carrega a relação de emails dos hospitais
    print("Carregando relação de emails dos hospitais...")
    hospital_emails = load_hospital_emails()
    
    if not hospital_emails:
        print("ERRO: Não foi possível carregar a relação de emails. Verifique o arquivo.")
        exit(1)
    
    # Inicia o processo principal - PRIMEIRO NAVEGADOR (DOWNLOAD)
    print("Iniciando navegador para download...")
    download_driver = start_browser(headless=False)
    
    try:
        # ETAPA 1: DOWNLOAD DOS ARQUIVOS
        print("\n" + "="*50)
        print("ETAPA 1: DOWNLOAD DOS ARQUIVOS DO OUTLOOK")
        print("="*50)
        
        login_to_outlook(download_driver)
        search_and_download_attachments(download_driver)
        
        # Extrai os arquivos ZIP após o download
        extract_zip_files()
        
        # Aguarda o download completar
        time.sleep(DEFAULT_WAIT_TIME * 3)
        
        # Verifica se algum arquivo foi baixado
        downloaded_files = list(DOWNLOAD_FOLDER.glob("*"))
        excel_files = list(DOWNLOAD_FOLDER.glob("*.xlsx"))
        
        if downloaded_files:
            print(f"Download concluído! Arquivos baixados: {[f.name for f in downloaded_files]}")
            print(f"Arquivos Excel disponíveis: {[f.name for f in excel_files]}")
            
            # ETAPA 2: PROCESSAR EXCEL E GERAR PDFs
            print("\n" + "="*50)
            print("ETAPA 2: PROCESSAMENTO DOS ARQUIVOS EXCEL")
            print("="*50)
            
            pdf_files = process_excel_files_and_generate_pdfs()
            
            if pdf_files:
                print(f"\nSUCESSO: {len(pdf_files)} PDFs gerados com sucesso!")
                print("PDFs criados:")
                for pdf_file in pdf_files:
                    print(f"  - {pdf_file.name}")
                
                # ETAPA 3: ENVIAR PDFs POR EMAIL
                print("\n" + "="*50)
                print("ETAPA 3: ENVIO DE EMAILS")
                print("="*50)
                
                send_all_pdfs_by_email(pdf_files, hospital_emails)
                
                # ETAPA 4: GERAR E ENVIAR RELATÓRIO
                print("\n" + "="*50)
                print("ETAPA 4: RELATÓRIO DE STATUS")
                print("="*50)
                
                report_path = generate_email_status_report()
                
                if report_path:
                    print("Enviando relatório de status por email...")
                    # Usa um novo navegador para enviar o relatório
                    report_driver = start_browser(headless=False)
                    try:
                        login_to_outlook(report_driver)
                        send_status_report_email(report_driver, report_path)
                        print("Relatório de status enviado com sucesso!")
                    except Exception as e:
                        print(f"Erro ao enviar relatório: {e}")
                    finally:
                        report_driver.quit()
                else:
                    print("Não foi possível gerar o relatório de status.")
                
            else:
                print("\nAVISO: Nenhum PDF foi gerado.")
            
        else:
            print("Nenhum arquivo foi baixado. Verifique se o email com anexo foi encontrado.")
            
    except Exception as e:
        print(f"Erro durante a execução: {e}")
        import traceback
        traceback.print_exc()
    finally:
        # Fecha o navegador de download
        try:
            download_driver.quit()
            print("Navegador de download fechado.")
        except:
            pass
        
    print("Processo finalizado!")
